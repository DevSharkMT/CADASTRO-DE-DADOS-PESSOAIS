import openpyxl
workbook = openpyxl.load_workbook('python.xlsx')
planilha = workbook['Usuários']
planilha = workbook.active
import os # = importa o limpa tela --> os.system('cls')
import time # = importa tempo de tela --> time.sleep()
from tabulate import tabulate 

def cadastro():
    print("=======================MENU DE CADASTRO=======================")
    nome= input("Digite o seu nome: ")
    cpf = input("Digite o seu cpf: ")
    tel = input("Digite o seu telefone: ")
    mail = input("Digite o seu email: ")
    ende = input("Digite o seu endereço: ")
    nasc = input("Digite seu ano de nascimento: ")
    valor_cadastro = len(planilha['G'])
    codigo = valor_cadastro
    nova_linha = [nome, cpf, tel, mail, ende, nasc, codigo]
    print("CADASTRO REALIZADO COM SUCESSO")

    planilha.append(nova_linha)
    workbook.save('python.xlsx')

def excluir():
    cod_excluir = (input("INFORME O CÓDIGO DE CADASTRO A SER EXCLUÍDO: "))
    for row in planilha.iter_rows(min_row=2, max_col=7):
        for cell in row:
            if (str(cell.value) == cod_excluir):
                planilha.delete_rows(cell.row)
                workbook.save('python.xlsx')

def listagem():
    print("=======================LISTAGEM=======================")
    cabec = [ # = responsável pelo cabeçalho da planilha
        'NOME',
        'CPF',
        'TELEFONE',
        'EMAIL',
        'ENDEREÇO',
        'NASCIMENTO'
    ]
    print("ola")
    dados = []
    i = 0
    for row in planilha.iter_rows(min_row=2, max_col=7,):
        dados.append([])
        for cell in row:
            dados[i].append(cell.value)
        i = i + 1
    print(tabulate(dados, headers= cabec, tablefmt="rounded_grid", stralign="center", numalign="center"))
# (values_only=True)

def atualizar():
    print("=======================MENU DE ATUALIZAÇÃO=======================")
    cod_atualizar = input("INFORME O CÓDIGO DE CADASTRO QUE DESEJA ATUALIZAR: ")
    for row in planilha.iter_rows(min_row=2, min_col=7,):
        for cell in row:
            if (str(cell.value) == cod_atualizar):
                linha = cell.row
                novo_nome = input("INFORME O NOVO NOME: ")
                novo_cpf = input("INFORME O NOVO CPF: ")
                novo_tel = input("INFORME O NOVO TELEFONE: ")
                novo_mail = input("INFORME O NOVO EMAIL: ")
                novo_ende = input("INFORME O NOVO ENDEREÇO: ")
                novo_nasc = input("INFORME A NOVA DATA DE NASCIMENTO: ")
                atualizacao = ["",novo_nome,novo_cpf,novo_tel,novo_mail,novo_ende,novo_nasc]
                for i in range(1,7):
                    planilha.cell(linha, i, atualizacao[i])
                    workbook.save('python.xlsx')

while True:
    print(
        "              SISTEMA DE CADASTROS              "
        "\n"
        "\nESCOLHA UMA DAS OPÇÕES ABAIXO:"
        "\n1) CADASTRAR"
        "\n2) LISTAR"
        "\n3) EXCLUIR"
        "\n4) ATUALIZAR"
        "\n0) SAIR DO SISTEMAS"
    )
    
    op = input("OPÇÃO DESEJADA --> ")
    
    if op == "1":
        cadastro()
    
    if op == "2":
        listagem()
    
    if op == "3":
        excluir()
    
    if op == "4":
        atualizar()

    if op == "0":
        print("SAINDO DO SISTEMA...")
        time.sleep(4)
        break

